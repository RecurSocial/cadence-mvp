#!/usr/bin/env python3
"""
Seed script: Import Euphoria Esthetics & Wellness data from Excel into Supabase
Reads /mnt/user-data/outputs/Euphoria_Service_Matrix.xlsx and populates:
- organizations
- services
- practitioners
- practitioner_certifications
"""

import os
import sys
from openpyxl import load_workbook
from supabase import create_client, Client

# Supabase config (from .env.local)
SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "https://nkjerngmimqkctmtecrc.supabase.co")
SUPABASE_KEY = os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5ramVybmdtaW1xa2N0bXRlY3JjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzU2ODMzOTcsImV4cCI6MjA5MTI1OTM5N30.iXATE_wIbdgfAkTNhqsUrFzLyrnF9LKss_yFciIxums")

# Euphoria org details
ORG_ID = "test-org-1"
ORG_NAME = "Euphoria Esthetics & Wellness"

# Practitioner mapping (name -> role, approval_level)
PRACTITIONERS_DATA = {
    "Brianna Krug": ("Nurse", "owner"),
    "Jaimie Burkett": ("Nurse", "staff"),
    "Kim Benitez": ("Nurse", "staff"),
    "Lexy Fazzone": ("Nurse", "staff"),
    "Michelle Wilson": ("Nurse", "owner"),
    "Nadine Delia": ("Nurse", "staff"),
    "Daisy": ("PA", "staff"),
    "Jordan Land": ("PA", "staff"),
    "Nicole Roberto": ("Aesthetician", "staff"),
    "Nicole Rekus": ("Aesthetician", "staff"),
    "Tori Grant": ("Aesthetician", "staff"),
    "Aubrey Rieger": ("Masseuse", "staff"),
}

def main():
    print("=" * 80)
    print(f"EUPHORIA SEED DATA IMPORT")
    print(f"Importing from: /mnt/user-data/outputs/Euphoria_Service_Matrix.xlsx")
    print(f"Org ID: {ORG_ID}")
    print("=" * 80)
    print()

    # Initialize Supabase client
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("✓ Connected to Supabase")
    print()

    # Step 1: Create/verify organization
    print("[1/4] Creating organization...")
    try:
        org_response = supabase.table("organizations").insert({
            "id": ORG_ID,
            "name": ORG_NAME,
            "slug": "euphoria-test",
            "plan_tier": "starter"
        }).execute()
        print(f"✓ Organization created: {ORG_NAME}")
    except Exception as e:
        if "duplicate" in str(e).lower():
            print(f"✓ Organization already exists: {ORG_NAME}")
        else:
            print(f"✗ Error: {e}")
            raise
    print()

    # Step 2: Load Excel and create practitioners
    print("[2/4] Creating practitioners...")
    wb = load_workbook('/mnt/user-data/outputs/Euphoria_Service_Matrix.xlsx', read_only=True)
    
    practitioners_map = {}  # name -> id
    for prac_name, (role, approval_level) in PRACTITIONERS_DATA.items():
        try:
            response = supabase.table("practitioners").insert({
                "org_id": ORG_ID,
                "first_name": prac_name.split()[0],
                "last_name": " ".join(prac_name.split()[1:]) if len(prac_name.split()) > 1 else "",
                "role": role,
                "approval_level": approval_level,
                "is_active": True,
            }).execute()
            
            prac_id = response.data[0]["id"]
            practitioners_map[prac_name] = prac_id
            print(f"  ✓ {prac_name} ({role}) - ID: {prac_id}")
        except Exception as e:
            print(f"  ✗ Error creating {prac_name}: {e}")
    print(f"✓ Created {len(practitioners_map)} practitioners")
    print()

    # Step 3: Process services and certifications
    print("[3/4] Creating services and certifications...")
    services_created = 0
    certifications_created = 0
    
    # Map of Excel column index to practitioner name
    header_row = None
    
    # Process each service category sheet
    for sheet_name in wb.sheetnames:
        if sheet_name == "INDEX":
            continue
        
        category = sheet_name
        ws = wb[sheet_name]
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # First row is header
            if row_idx == 1:
                header_row = row
                continue
            
            # Skip empty rows
            if not row[0]:
                break
            
            service_name = row[0]
            product = row[1] or "N/A"
            supplier = row[2] or "N/A"
            duration_minutes = row[3]
            price = str(row[4]) if row[4] else "Contact"
            
            # Create service
            try:
                service_response = supabase.table("services").insert({
                    "org_id": ORG_ID,
                    "category": category,
                    "name": service_name,
                    "product": product,
                    "supplier": supplier,
                    "duration_minutes": duration_minutes,
                    "price": price,
                }).execute()
                
                service_id = service_response.data[0]["id"]
                services_created += 1
                
                # Extract certifications from columns 5+ (practitioner columns)
                # Column mapping: 5->Brianna, 6->Jaimie, 7->Kim, etc.
                for col_idx in range(5, len(header_row)):
                    if col_idx >= len(row):
                        break
                    
                    prac_name = header_row[col_idx]
                    certification_value = row[col_idx]
                    
                    # Check if certified (☑ or True)
                    is_certified = certification_value == "☑"
                    
                    if prac_name in practitioners_map:
                        prac_id = practitioners_map[prac_name]
                        try:
                            supabase.table("practitioner_certifications").insert({
                                "practitioner_id": prac_id,
                                "service_id": service_id,
                                "certified": is_certified,
                            }).execute()
                            if is_certified:
                                certifications_created += 1
                        except Exception as cert_error:
                            if "duplicate" not in str(cert_error).lower():
                                print(f"    ✗ Certification error for {prac_name}: {cert_error}")
            
            except Exception as e:
                print(f"  ✗ Error creating service '{service_name}': {e}")
    
    print(f"✓ Created {services_created} services with {certifications_created} certifications")
    print()

    # Step 4: Verify data
    print("[4/4] Verifying data...")
    try:
        services_count = supabase.table("services").select("id", count="exact").eq("org_id", ORG_ID).execute()
        practitioners_count = supabase.table("practitioners").select("id", count="exact").eq("org_id", ORG_ID).execute()
        
        print(f"✓ Total services: {len(services_count.data)}")
        print(f"✓ Total practitioners: {len(practitioners_count.data)}")
    except Exception as e:
        print(f"✗ Error verifying: {e}")
    
    print()
    print("=" * 80)
    print("✓ SEED COMPLETE!")
    print("=" * 80)
    print()
    print("Next step: Reload https://cadence-mvp.vercel.app/dashboard")
    print()

if __name__ == "__main__":
    main()
